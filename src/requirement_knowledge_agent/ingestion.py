from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .io import read_json
from .io import write_json
from .validation import KnowledgeValidationError, validate_default_solution
from .validation import validate_standard_clause


LIST_FIELDS = {
    "trigger_terms",
    "boundary_conditions",
    "acceptance_criteria",
    "confirmation_questions",
    "related_standard_clause_ids",
}
REQUIRED_SOLUTION_COLUMNS = (
    "solution_id",
    "module",
    "submodule",
    "scenario",
    "trigger_terms",
    "default_behavior",
)
STANDARD_LIST_FIELDS = {"keywords", "applies_to"}
METER_REQUIREMENT_SHEETS = {
    "系统需求": "SYSTEM",
    "计量需求": "MEASUREMENT",
    "时钟需求": "CLOCK",
    "费率需求": "TARIFF",
    "显示需求": "DISPLAY",
    "需量需求": "DEMAND",
    "结算需求": "BILLING",
    "负荷曲线": "LOAD-PROFILE",
    "报警窃电需求": "ALARM-TAMPER",
    "电网质量需求": "POWER-QUALITY",
    "升级需求": "UPGRADE",
    "负控需求": "LOAD-CONTROL",
    "状态字需求": "STATUS-WORD",
    "事件需求": "EVENT",
    "协议栈需求": "PROTOCOL",
    "push需求": "PUSH",
    "P1需求": "P1",
    "MBUS需求": "MBUS",
    "预付费需求": "PREPAYMENT",
}
METER_REQUIRED_HEADERS = {"子模块", "描述", "说明、示例、注意事项"}
METER_CONFIRMATION_TERMS = (
    "需明确",
    "请填写",
    "如支持",
    "如可配置",
    "取决于项目",
    "需说明",
    "确认",
    "根据客户需求",
)


def ingest_solutions_excel(input_path: Path, out_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, data_only=True)
    sheet = workbook.active
    headers = _read_headers(sheet)
    payload: list[dict[str, Any]] = []
    issues: list[str] = []
    for row_number, values in _iter_data_rows(sheet, headers):
        try:
            raw = _solution_from_row(headers, values)
        except KnowledgeValidationError as exc:
            issues.extend(f"row {row_number}: {issue}" for issue in exc.issues)
            continue
        validation_issues = validate_default_solution(raw)
        issues.extend(f"row {row_number}: {issue}" for issue in validation_issues)
        payload.append(raw)
    if issues:
        raise KnowledgeValidationError(issues)
    write_json(out_path, payload)
    return payload


def ingest_meter_template_solutions(input_path: Path, out_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, data_only=True)
    payload: list[dict[str, Any]] = []
    issues: list[str] = []
    supported_sheets = 0
    for sheet in workbook.worksheets:
        if sheet.title not in METER_REQUIREMENT_SHEETS:
            continue
        headers = [_cell_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        header_map = {header: index for index, header in enumerate(headers) if header}
        if not _is_meter_requirement_sheet(header_map):
            continue
        supported_sheets += 1
        for row_number, values in _iter_meter_template_rows(sheet, headers):
            raw = _meter_solution_from_row(sheet.title, row_number, header_map, values)
            if raw is None:
                continue
            validation_issues = validate_default_solution(raw)
            issues.extend(f"{sheet.title} row {row_number}: {issue}" for issue in validation_issues)
            payload.append(raw)
    if issues:
        raise KnowledgeValidationError(issues)
    if supported_sheets == 0:
        raise KnowledgeValidationError(["no supported meter template requirement sheets found"])
    if not payload:
        raise KnowledgeValidationError(["no meter template default solutions found"])
    write_json(out_path, payload)
    return payload


def ingest_standards(input_path: Path, out_path: Path) -> list[dict[str, Any]]:
    payload: list[dict[str, Any]] = []
    issues: list[str] = []
    for path in _standard_input_files(input_path):
        try:
            raw_items = _read_standard_items(path)
        except KnowledgeValidationError as exc:
            issues.extend(f"{path.name}: {issue}" for issue in exc.issues)
            continue
        for raw in raw_items:
            validation_issues = validate_standard_clause(raw)
            issues.extend(f"{path.name}: {issue}" for issue in validation_issues)
            payload.append(raw)
    if issues:
        raise KnowledgeValidationError(issues)
    write_json(out_path, payload)
    return payload


def _standard_input_files(input_path: Path) -> list[Path]:
    if input_path.is_dir():
        return sorted(
            path
            for path in input_path.iterdir()
            if path.is_file() and path.suffix.lower() in {".md", ".json"}
        )
    return [input_path]


def _read_standard_items(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".json":
        payload = read_json(path)
        if isinstance(payload, list):
            return payload
        if isinstance(payload, dict) and isinstance(payload.get("items"), list):
            return payload["items"]
        raise KnowledgeValidationError(["JSON must be a list or an object with an items list"])
    if path.suffix.lower() == ".md":
        return [_standard_from_markdown(path)]
    raise KnowledgeValidationError(["input must be a .md file, .json file, or directory"])


def _standard_from_markdown(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8-sig")
    metadata, body = _split_frontmatter(text)
    raw: dict[str, Any] = {}
    for key, value in metadata.items():
        if key in STANDARD_LIST_FIELDS:
            raw[key] = _parse_list(value)
        else:
            raw[key] = value
    raw.setdefault("source_file", path.name)
    raw.setdefault("constraint_level", "reference")
    raw["text"] = body.strip()
    if raw.get("source_section") and not raw.get("citation"):
        raw["citation"] = f"{raw['source_file']} section {raw['source_section']}"
    return raw


def _split_frontmatter(text: str) -> tuple[dict[str, str], str]:
    if not text.startswith("---"):
        raise KnowledgeValidationError(["markdown frontmatter is required"])
    lines = text.splitlines()
    end_index = None
    for index, line in enumerate(lines[1:], start=1):
        if line.strip() == "---":
            end_index = index
            break
    if end_index is None:
        raise KnowledgeValidationError(["markdown frontmatter closing marker is required"])
    metadata: dict[str, str] = {}
    for line in lines[1:end_index]:
        if not line.strip():
            continue
        if ":" not in line:
            raise KnowledgeValidationError([f"invalid frontmatter line: {line}"])
        key, value = line.split(":", 1)
        metadata[key.strip()] = value.strip()
    return metadata, "\n".join(lines[end_index + 1 :])


def _read_headers(sheet) -> list[str]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        raise KnowledgeValidationError(["row 1: header row is required"])
    headers = [_normalize_header(value) for value in first_row]
    missing = [field for field in REQUIRED_SOLUTION_COLUMNS if field not in headers]
    if missing:
        raise KnowledgeValidationError([f"missing required column: {field}" for field in missing])
    return headers


def _iter_data_rows(sheet, headers: list[str]):
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[: len(headers)])
        if any(_cell_text(value) for value in values):
            yield row_number, values


def _solution_from_row(headers: list[str], values: list[Any]) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for header, value in zip(headers, values):
        if not header:
            continue
        if header in LIST_FIELDS:
            raw[header] = _parse_list(value)
        elif header == "config_items":
            raw[header] = _parse_config_items(value)
        elif header == "requires_confirmation":
            raw[header] = _parse_bool(value)
        else:
            raw[header] = _cell_text(value)
    for field in LIST_FIELDS:
        raw.setdefault(field, [])
    raw.setdefault("config_items", [])
    raw.setdefault("requires_confirmation", False)
    return raw


def _is_meter_requirement_sheet(header_map: dict[str, int]) -> bool:
    return METER_REQUIRED_HEADERS.issubset(header_map)


def _iter_meter_template_rows(sheet, headers: list[str]):
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[: len(headers)])
        if any(_cell_text(value) for value in values):
            yield row_number, values


def _meter_solution_from_row(
    sheet_name: str,
    row_number: int,
    header_map: dict[str, int],
    values: list[Any],
) -> dict[str, Any] | None:
    if _meter_cell("关闭", header_map, values) in {"是", "true", "1", "y", "yes", "关闭"}:
        return None
    submodule = _meter_cell("子模块", header_map, values)
    scenario = _meter_cell("描述", header_map, values)
    requirement = _meter_cell("需求", header_map, values)
    template = _meter_cell("需求模版", header_map, values)
    notes = _meter_cell("说明、示例、注意事项", header_map, values)
    default_behavior = requirement or template or notes
    if not (submodule and scenario and default_behavior):
        return None
    requires_confirmation = _requires_meter_confirmation(requirement, template, notes)
    return {
        "solution_id": _meter_solution_id(sheet_name, row_number),
        "module": sheet_name,
        "submodule": submodule,
        "scenario": scenario,
        "trigger_terms": _meter_trigger_terms(sheet_name, submodule, scenario),
        "default_behavior": default_behavior,
        "config_items": [],
        "boundary_conditions": _meter_boundary_conditions(notes),
        "acceptance_criteria": _meter_acceptance_criteria(scenario, default_behavior),
        "confirmation_questions": _meter_confirmation_questions(submodule, scenario, notes, requires_confirmation),
        "related_standard_clause_ids": [],
        "requires_confirmation": requires_confirmation,
    }


def _meter_cell(header: str, header_map: dict[str, int], values: list[Any]) -> str:
    if header not in header_map:
        return ""
    index = header_map[header]
    if index >= len(values):
        return ""
    return _cell_text(values[index])


def _meter_solution_id(sheet_name: str, row_number: int) -> str:
    sheet_code = METER_REQUIREMENT_SHEETS[sheet_name]
    return f"SOL-METER-{sheet_code}-{row_number:04d}"


def _meter_trigger_terms(sheet_name: str, submodule: str, scenario: str) -> list[str]:
    terms = [sheet_name, _sheet_topic(sheet_name), submodule, _clean_scenario(scenario)]
    return _dedupe_texts(terms)


def _sheet_topic(sheet_name: str) -> str:
    return sheet_name[:-2] if sheet_name.endswith("需求") else sheet_name


def _clean_scenario(scenario: str) -> str:
    return re.sub(r"[:：\\s]+$", "", scenario).strip()


def _meter_boundary_conditions(notes: str) -> list[str]:
    return _split_meter_notes(notes)


def _meter_acceptance_criteria(scenario: str, default_behavior: str) -> list[str]:
    clean_scenario = _clean_scenario(scenario)
    if clean_scenario:
        return [f"{clean_scenario}按默认方案实现：{default_behavior}"]
    return [f"按默认方案实现：{default_behavior}"]


def _meter_confirmation_questions(
    submodule: str,
    scenario: str,
    notes: str,
    requires_confirmation: bool,
) -> list[str]:
    if not requires_confirmation or not notes:
        return []
    questions = []
    for item in _split_meter_notes(notes):
        questions.append(f"请确认{submodule}/{scenario}{item}")
    return questions


def _requires_meter_confirmation(*texts: str) -> bool:
    joined = "\n".join(text for text in texts if text)
    return any(term in joined for term in METER_CONFIRMATION_TERMS)


def _split_meter_notes(text: str) -> list[str]:
    if not text:
        return []
    lines = []
    for line in re.split(r"[\n;；]+", text):
        clean = line.strip()
        if clean:
            lines.append(clean)
    return _dedupe_texts(lines)


def _dedupe_texts(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        clean = _cell_text(value)
        if clean and clean not in seen:
            seen.add(clean)
            result.append(clean)
    return result


def _normalize_header(value: Any) -> str:
    return _cell_text(value).strip().lower()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_list(value: Any) -> list[str]:
    text = _cell_text(value)
    if not text:
        return []
    return [item.strip() for item in re.split("[;\\n\\uFF1B]+", text) if item.strip()]


def _parse_config_items(value: Any) -> list[dict[str, Any]]:
    text = _cell_text(value)
    if not text:
        return []
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise KnowledgeValidationError(["config_items must be valid JSON"]) from exc
    if isinstance(payload, dict):
        return [payload]
    if isinstance(payload, list) and all(isinstance(item, dict) for item in payload):
        return payload
    raise KnowledgeValidationError(["config_items must be a JSON object or array of objects"])


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = _cell_text(value).lower()
    return text in {"1", "true", "yes", "y", "\u662f", "\u9700\u8981", "\u9700\u786e\u8ba4"}
