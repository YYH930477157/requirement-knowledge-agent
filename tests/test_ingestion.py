import json

import pytest
from openpyxl import Workbook

import requirement_knowledge_agent.ingestion as ingestion
from requirement_knowledge_agent.ingestion import ingest_solutions_excel, ingest_standards
from requirement_knowledge_agent.validation import KnowledgeValidationError


def write_workbook(path, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def write_meter_template_workbook(path):
    workbook = Workbook()
    system = workbook.active
    system.title = "系统需求"
    system.append(
        [
            "关闭",
            "序号",
            "子模块",
            "描述",
            "需求模版",
            "需求",
            "说明、示例、注意事项",
            "是否客户需求",
        ]
    )
    system.append(["", "1", "基本参数", "电表类型：", "1P2W_SP", "3P4W_DC", "如可配置，需说明配置范围", "", ""])

    measurement = workbook.create_sheet("计量需求")
    measurement.append(["关闭", "序号", "子模块", "描述", "1P2W_SP", "说明、示例、注意事项"])
    measurement.append(["", "1", "基本参数", "火线采样类型：", "1", "CT采样；锰铜采样；无"])

    event = workbook.create_sheet("事件需求")
    event.append(["关闭", "序号", "子模块", "描述", "需求", "说明、示例、注意事项"])
    event.append(["", "1", "事件", "事件log filter：", "支持", "建议都实现该功能，以防止现场的特殊情况出现"])

    list_sheet = workbook.create_sheet("计量列表")
    list_sheet.append(["瞬时数据列表", "OBIS CODE", "描述"])
    list_sheet.append(["1", "{3,1-0:32.7.0.255},", "Instantaneous Voltage of L1"])

    workbook.save(path)


def test_ingest_solutions_excel_writes_runtime_json(tmp_path):
    source = tmp_path / "default_solutions.xlsx"
    out = tmp_path / "kb" / "default_solutions.json"
    write_workbook(
        source,
        [
            "solution_id",
            "module",
            "submodule",
            "scenario",
            "trigger_terms",
            "default_behavior",
            "config_items",
            "boundary_conditions",
            "acceptance_criteria",
            "confirmation_questions",
            "related_standard_clause_ids",
            "requires_confirmation",
        ],
        [
            [
                "SOL-DISPLAY-1",
                "Display",
                "Cycle",
                "Automatic display cycle",
                "display; cycle\nLCD；screen",
                "Cycle configured display entries.",
                '[{"name":"cycle_interval_seconds","default_value":"5","requires_confirmation":true}]',
                "No entries means no cycle.\nManual override pauses cycling.",
                "Entries are shown in order; Cycle interval can be configured.",
                "Confirm display list source.\nConfirm cycle interval.",
                "STD-DISPLAY-1; STD-DISPLAY-2",
                "yes",
            ]
        ],
    )

    payload = ingest_solutions_excel(source, out)

    assert out.exists()
    assert payload[0]["trigger_terms"] == ["display", "cycle", "LCD", "screen"]
    assert payload[0]["boundary_conditions"] == [
        "No entries means no cycle.",
        "Manual override pauses cycling.",
    ]
    assert payload[0]["acceptance_criteria"] == [
        "Entries are shown in order",
        "Cycle interval can be configured.",
    ]
    assert payload[0]["related_standard_clause_ids"] == ["STD-DISPLAY-1", "STD-DISPLAY-2"]
    assert payload[0]["confirmation_questions"] == [
        "Confirm display list source.",
        "Confirm cycle interval.",
    ]
    assert payload[0]["config_items"] == [
        {
            "name": "cycle_interval_seconds",
            "default_value": "5",
            "requires_confirmation": True,
        }
    ]
    assert payload[0]["requires_confirmation"] is True
    assert json.loads(out.read_text(encoding="utf-8")) == payload


def test_ingest_meter_template_solutions_converts_requirement_sheets(tmp_path):
    source = tmp_path / "meter_template.xlsx"
    out = tmp_path / "kb" / "default_solutions.json"
    write_meter_template_workbook(source)

    payload = ingestion.ingest_meter_template_solutions(source, out)

    assert out.exists()
    assert [item["solution_id"] for item in payload] == [
        "SOL-METER-SYSTEM-0002",
        "SOL-METER-MEASUREMENT-0002",
        "SOL-METER-EVENT-0002",
    ]
    assert payload[0]["module"] == "系统需求"
    assert payload[0]["submodule"] == "基本参数"
    assert payload[0]["scenario"] == "电表类型："
    assert payload[0]["trigger_terms"] == ["系统需求", "系统", "基本参数", "电表类型"]
    assert payload[0]["default_behavior"] == "3P4W_DC"
    assert payload[0]["boundary_conditions"] == ["如可配置，需说明配置范围"]
    assert payload[0]["requires_confirmation"] is True
    assert payload[0]["confirmation_questions"] == ["请确认基本参数/电表类型：如可配置，需说明配置范围"]
    assert payload[0]["related_standard_clause_ids"] == []
    assert payload[1]["default_behavior"] == "CT采样；锰铜采样；无"
    assert payload[2]["default_behavior"] == "支持"
    assert json.loads(out.read_text(encoding="utf-8")) == payload


def test_ingest_meter_template_solutions_skips_closed_and_empty_rows(tmp_path):
    source = tmp_path / "meter_template.xlsx"
    out = tmp_path / "default_solutions.json"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "显示需求"
    sheet.append(["关闭", "序号", "子模块", "描述", "需求模版", "需求", "说明、示例、注意事项"])
    sheet.append(["是", "1", "显示格式", "显示溢出处理：", "多屏显示", "多屏显示", "关闭行不摄入"])
    sheet.append(["", "2", "显示格式", "显示轮显间隔：", "", "", ""])
    sheet.append(["", "3", "显示格式", "LCD背光：", "支持", "支持", "确认是否支持背光"])
    workbook.save(source)

    payload = ingestion.ingest_meter_template_solutions(source, out)

    assert [item["solution_id"] for item in payload] == ["SOL-METER-DISPLAY-0004"]
    assert payload[0]["scenario"] == "LCD背光："


def test_ingest_meter_template_solutions_reports_when_no_supported_sheet(tmp_path):
    source = tmp_path / "meter_template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "计量列表"
    sheet.append(["瞬时数据列表", "OBIS CODE", "描述"])
    sheet.append(["1", "{3,1-0:32.7.0.255},", "Instantaneous Voltage of L1"])
    workbook.save(source)

    with pytest.raises(KnowledgeValidationError) as excinfo:
        ingestion.ingest_meter_template_solutions(source, tmp_path / "default_solutions.json")

    assert "no supported meter template requirement sheets found" in str(excinfo.value)


def test_ingest_solutions_excel_reports_required_field_with_row_number(tmp_path):
    source = tmp_path / "default_solutions.xlsx"
    write_workbook(
        source,
        ["solution_id", "module", "submodule", "scenario", "trigger_terms", "default_behavior"],
        [["", "Display", "Cycle", "Automatic display cycle", "display", "Cycle entries."]],
    )

    with pytest.raises(KnowledgeValidationError) as excinfo:
        ingest_solutions_excel(source, tmp_path / "default_solutions.json")

    assert "row 2: solution_id is required" in str(excinfo.value)


def test_ingest_solutions_excel_reports_invalid_config_json(tmp_path):
    source = tmp_path / "default_solutions.xlsx"
    write_workbook(
        source,
        ["solution_id", "module", "submodule", "scenario", "trigger_terms", "default_behavior", "config_items"],
        [["SOL-1", "Display", "Cycle", "Automatic display cycle", "display", "Cycle entries.", "{bad json"]],
    )

    with pytest.raises(KnowledgeValidationError) as excinfo:
        ingest_solutions_excel(source, tmp_path / "default_solutions.json")

    assert "row 2: config_items must be valid JSON" in str(excinfo.value)


def test_ingest_standards_reads_markdown_directory(tmp_path):
    source = tmp_path / "standards"
    source.mkdir()
    (source / "display.md").write_text(
        """---
clause_id: STD-DISPLAY-1
source_section: 4.2.1
title: Display cycle behavior
keywords: display; cycle
applies_to: display
constraint_level: must
---
The device shall support configured display cycling.
""",
        encoding="utf-8",
    )
    out = tmp_path / "kb" / "standards.json"

    payload = ingest_standards(source, out)

    assert out.exists()
    assert payload == [
        {
            "clause_id": "STD-DISPLAY-1",
            "source_file": "display.md",
            "source_section": "4.2.1",
            "title": "Display cycle behavior",
            "text": "The device shall support configured display cycling.",
            "keywords": ["display", "cycle"],
            "applies_to": ["display"],
            "constraint_level": "must",
            "citation": "display.md section 4.2.1",
        }
    ]


def test_ingest_standards_reads_json_file(tmp_path):
    source = tmp_path / "standards.json"
    source.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "clause_id": "STD-1",
                        "source_file": "standard.json",
                        "source_section": "1",
                        "title": "Display",
                        "text": "The display shall cycle.",
                        "keywords": ["display"],
                        "applies_to": ["display"],
                        "constraint_level": "should",
                        "citation": "standard.json section 1",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    payload = ingest_standards(source, tmp_path / "kb" / "standards.json")

    assert payload[0]["clause_id"] == "STD-1"
    assert payload[0]["constraint_level"] == "should"


def test_ingest_standards_reports_invalid_constraint_with_file_name(tmp_path):
    source = tmp_path / "standards"
    source.mkdir()
    (source / "bad.md").write_text(
        """---
clause_id: STD-BAD
source_section: 1
title: Bad clause
keywords: bad
applies_to: display
constraint_level: optional
---
Bad clause text.
""",
        encoding="utf-8",
    )

    with pytest.raises(KnowledgeValidationError) as excinfo:
        ingest_standards(source, tmp_path / "standards.json")

    assert "bad.md: constraint_level must be one of: must, should, reference" in str(excinfo.value)
