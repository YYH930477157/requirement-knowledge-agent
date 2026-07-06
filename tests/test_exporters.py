import json

from openpyxl import load_workbook

from requirement_knowledge_agent.exporters import export_review_package


def package():
    return {
        "summary": {"total_requirements": 1, "analyzed": 1, "input_errors": 0, "decisions": {"applied": 1}},
        "items": [
            {
                "requirement_id": "REQ-1",
                "source_text": "需要显示轮显",
                "module": "显示",
                "submodule": "轮显",
                "decision": "applied",
                "confidence": 0.9,
                "landing_requirement": "软件应按配置顺序轮显条目。",
                "developer_guidance": ["按配置顺序轮显。"],
                "acceptance_criteria": ["按顺序显示。"],
                "applied_solution_ids": ["SOL-1"],
                "standard_citations": [{"clause_id": "STD-1", "citation": "s.md 1", "constraint_level": "must"}],
                "open_questions": [],
                "reasoning_summary": "强命中。",
            }
        ],
        "input_errors": [],
    }


def test_export_review_package_writes_json(tmp_path):
    export_review_package(package(), tmp_path)
    payload = json.loads((tmp_path / "review_package.json").read_text(encoding="utf-8"))
    assert payload["items"][0]["requirement_id"] == "REQ-1"


def test_markdown_groups_by_decision(tmp_path):
    export_review_package(package(), tmp_path)
    text = (tmp_path / "review_package.md").read_text(encoding="utf-8")
    assert "## applied" in text
    assert "REQ-1" in text


def test_excel_export_writes_headers_and_row(tmp_path):
    export_review_package(package(), tmp_path)
    wb = load_workbook(tmp_path / "software_requirements.xlsx")
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "需求编号"
    assert ws.cell(row=2, column=1).value == "REQ-1"
